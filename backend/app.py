import io
import os
import uuid
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from flask import Flask, jsonify, request, g, send_file
from flask_cors import CORS
from dotenv import load_dotenv

import auth
import column_mapper
import data_adapter as da
import intent_engine
import query_engine
import response_builder

load_dotenv()

# ── Excel template definition ─────────────────────────────────────────────────

_TEMPLATE_COLUMNS = [
    "Product Code", "Product Name", "Year",
    "Opening Stock", "Quantity Purchased", "Closing Stock", "Sales Channel",
]

_SAMPLE_DATA = [
    {"Product Code": "SKI-JKT-001", "Product Name": "Ski Jacket Pro",    "Sales Channel": "Main Store", "Year": 2022, "Opening Stock": 50,  "Quantity Purchased": 300, "Closing Stock": 40},
    {"Product Code": "SKI-JKT-001", "Product Name": "Ski Jacket Pro",    "Sales Channel": "Main Store", "Year": 2023, "Opening Stock": 40,  "Quantity Purchased": 320, "Closing Stock": 35},
    {"Product Code": "SKI-JKT-001", "Product Name": "Ski Jacket Pro",    "Sales Channel": "Main Store", "Year": 2024, "Opening Stock": 35,  "Quantity Purchased": 350, "Closing Stock": 30},
    {"Product Code": "SKI-JKT-001", "Product Name": "Ski Jacket Pro",    "Sales Channel": "Online",     "Year": 2022, "Opening Stock": 10,  "Quantity Purchased":  80, "Closing Stock":  8},
    {"Product Code": "SKI-JKT-001", "Product Name": "Ski Jacket Pro",    "Sales Channel": "Online",     "Year": 2023, "Opening Stock":  8,  "Quantity Purchased": 100, "Closing Stock":  6},
    {"Product Code": "SKI-JKT-001", "Product Name": "Ski Jacket Pro",    "Sales Channel": "Online",     "Year": 2024, "Opening Stock":  6,  "Quantity Purchased": 130, "Closing Stock":  5},
    {"Product Code": "SKI-BT-002",  "Product Name": "Alpine Ski Boots",  "Sales Channel": "Main Store", "Year": 2022, "Opening Stock": 80,  "Quantity Purchased": 200, "Closing Stock": 60},
    {"Product Code": "SKI-BT-002",  "Product Name": "Alpine Ski Boots",  "Sales Channel": "Main Store", "Year": 2023, "Opening Stock": 60,  "Quantity Purchased": 210, "Closing Stock": 50},
    {"Product Code": "SKI-BT-002",  "Product Name": "Alpine Ski Boots",  "Sales Channel": "Main Store", "Year": 2024, "Opening Stock": 50,  "Quantity Purchased": 230, "Closing Stock": 45},
    {"Product Code": "GLOVES-003",  "Product Name": "Thermal Gloves",    "Sales Channel": "Main Store", "Year": 2022, "Opening Stock": 120, "Quantity Purchased": 500, "Closing Stock": 80},
    {"Product Code": "GLOVES-003",  "Product Name": "Thermal Gloves",    "Sales Channel": "Main Store", "Year": 2023, "Opening Stock": 80,  "Quantity Purchased": 540, "Closing Stock": 70},
    {"Product Code": "GLOVES-003",  "Product Name": "Thermal Gloves",    "Sales Channel": "Main Store", "Year": 2024, "Opening Stock": 70,  "Quantity Purchased": 580, "Closing Stock": 60},
]


def _build_template_excel() -> bytes:
    """Generate and return the Excel planning template as bytes."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Planning Template"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    example_fill = PatternFill("solid", fgColor="EBF3FB")
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    col_widths = [18, 22, 8, 16, 20, 16, 18]

    for col_idx, (col_name, width) in enumerate(zip(_TEMPLATE_COLUMNS, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 28

    for row_idx, row_data in enumerate(_SAMPLE_DATA, start=2):
        for col_idx, col_name in enumerate(_TEMPLATE_COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ""))
            cell.fill = example_fill
            cell.border = border
            cell.alignment = center

    ws_info = wb.create_sheet("Instructions")
    ws_info.column_dimensions["A"].width = 80
    instructions = [
        ("Planning Tool — Excel Template Instructions", True),
        ("", False),
        ("1. Do NOT rename or remove any column headers in the 'Planning Template' sheet.", False),
        ("2. Delete the example rows (rows 2 onwards) and enter your own data.", False),
        ("3. Each row = ONE product, ONE year, ONE sales channel (optional).", False),
        ("4. Required columns: Product Code, Product Name, Year, Opening Stock, Quantity Purchased, Closing Stock.", False),
        ("5. Optional: Sales Channel — leave blank if you have only one channel.", False),
        ("6. Sales = Opening Stock + Quantity Purchased − Closing Stock.", False),
        ("7. Include at least 2-3 years of data per product for better forecasts.", False),
    ]
    for r_idx, (text, bold) in enumerate(instructions, start=1):
        cell = ws_info.cell(row=r_idx, column=1, value=text)
        if bold:
            cell.font = Font(bold=True, size=13)
        ws_info.row_dimensions[r_idx].height = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


app = Flask(__name__)
CORS(app)

auth.init_db()

# In-memory session store: {session_id: {"adapter": InMemoryAdapter, "meta": {...}}}
_sessions: dict = {}


# ── Auth ──────────────────────────────────────────────────────────────────────

@app.route("/api/auth/register", methods=["POST"])
def register():
    body = request.get_json(silent=True) or {}
    email    = str(body.get("email", "")).strip()
    password = str(body.get("password", "")).strip()

    if not email or "@" not in email:
        return jsonify({"error": "Indirizzo email non valido."}), 400
    if len(password) < 6:
        return jsonify({"error": "La password deve contenere almeno 6 caratteri."}), 400

    user = auth.create_user(email, password)
    if user is None:
        return jsonify({"error": "Email già registrata."}), 409

    token = auth.create_token(user["id"])
    return jsonify({"token": token, "user": user}), 201


@app.route("/api/auth/login", methods=["POST"])
def login():
    body = request.get_json(silent=True) or {}
    email    = str(body.get("email", "")).strip()
    password = str(body.get("password", "")).strip()

    user = auth.verify_user(email, password)
    if not user:
        return jsonify({"error": "Email o password non corretti."}), 401

    token = auth.create_token(user["id"])
    return jsonify({"token": token, "user": user})


@app.route("/api/auth/me", methods=["GET"])
@auth.require_auth
def me():
    return jsonify({"user_id": g.user_id})


# ── Excel parsing ─────────────────────────────────────────────────────────────

@app.route("/api/parse-excel", methods=["POST"])
@auth.require_auth
def parse_excel():
    if "file" not in request.files:
        return jsonify({"error": "Nessun file ricevuto. Campo atteso: 'file'."}), 400

    file = request.files["file"]
    if not file.filename:
        return jsonify({"error": "Nessun file selezionato."}), 400
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"error": "Sono supportati solo file Excel (.xlsx, .xls)."}), 400

    contents = file.read()

    # ── Detect multi-sheet report format (Acquistato/Venduto per stagione) ──────
    try:
        xl_sheets = pd.ExcelFile(io.BytesIO(contents)).sheet_names
    except Exception:
        xl_sheets = []

    if column_mapper.is_multisheet_report(xl_sheets):
        products, category, err = column_mapper.parse_multisheet_report(contents)
        if err:
            return jsonify({"error": err}), 422

        adapter = da.InMemoryAdapter(products)
        session_id = str(uuid.uuid4())
        _sessions[session_id] = {
            "adapter":  adapter,
            "meta": {
                "categories":    adapter.get_categories(),
                "attributes":    adapter.get_attributes(),
                "years":         adapter.get_years(),
                "suppliers":     adapter.get_suppliers(),
                "product_count": len({p["product_code"] for p in products}),
            },
            "user_id":    g.user_id,
            "created_at": datetime.utcnow().isoformat(),
        }
        meta = _sessions[session_id]["meta"]
        return jsonify({
            "session_id":       session_id,
            "needs_confirmation": False,
            "product_count":    meta["product_count"],
            "categories":       meta["categories"],
            "attributes":       meta["attributes"],
            "years":            meta["years"],
        })

    # ── Standard single-sheet / multi-header formats ──────────────────────────
    try:
        df = pd.read_excel(io.BytesIO(contents))
    except Exception as exc:
        return jsonify({"error": f"Impossibile leggere il file Excel: {str(exc)}"}), 422

    # Detect two-row headers (many Unnamed columns = Atelier-style season exports)
    unnamed_count = sum(1 for c in df.columns if "Unnamed" in str(c))
    if unnamed_count >= 2:
        try:
            df_multi = pd.read_excel(io.BytesIO(contents), header=[0, 1])
            df_flat, was_multi = column_mapper.flatten_multiheader(df_multi)
            if was_multi:
                df = df_flat
        except Exception:
            pass  # fall through to single-header parsing

    headers = [str(c).strip() for c in df.columns.tolist()]
    sample_row = {}
    if not df.empty:
        first = df.iloc[0]
        sample_row = {str(c).strip(): first[c] for c in df.columns}

    mapping_result = column_mapper.map_columns(headers, sample_row)

    # Build a preview of the first 5 rows with original column names
    preview = df.head(5).fillna("").to_dict(orient="records")
    preview = [{str(k).strip(): str(v) for k, v in row.items()} for row in preview]

    # Temporarily store the raw df bytes for the confirm step
    session_id = str(uuid.uuid4())
    _sessions[session_id] = {
        "_pending_excel_bytes": contents,
        "adapter": None,
        "meta": {},
        "user_id": g.user_id,
        "created_at": datetime.utcnow().isoformat(),
    }

    response_year_groups = {str(k): v for k, v in mapping_result["year_groups"].items()}

    return jsonify({
        "session_id":      session_id,
        "mapping":         mapping_result["mapping"],
        "confidence":      mapping_result["confidence"],
        "year_groups":     response_year_groups,
        "unmapped":        mapping_result["unmapped"],
        "missing_required": mapping_result["missing_required"],
        "needs_confirmation": mapping_result["needs_confirmation"],
        "preview":         preview,
        "headers":         headers,
    })


@app.route("/api/parse-excel/confirm", methods=["POST"])
@auth.require_auth
def parse_excel_confirm():
    body = request.get_json(silent=True) or {}
    session_id = str(body.get("session_id", ""))
    mapping    = body.get("mapping", {})      # user-confirmed mapping
    year_groups_raw = body.get("year_groups", {})

    session = _sessions.get(session_id)
    if not session:
        return jsonify({"error": "Sessione non trovata o scaduta. Ricarica il file."}), 404
    if session.get("user_id") != g.user_id:
        return jsonify({"error": "Non autorizzato."}), 403

    excel_bytes = session.get("_pending_excel_bytes")
    if not excel_bytes:
        return jsonify({"error": "Dati Excel non trovati. Ricarica il file."}), 404

    try:
        df = pd.read_excel(io.BytesIO(excel_bytes))
    except Exception as exc:
        return jsonify({"error": f"Errore nella lettura del file: {str(exc)}"}), 422

    # Apply same multiheader flattening as in parse step
    unnamed_count = sum(1 for c in df.columns if "Unnamed" in str(c))
    if unnamed_count >= 2:
        try:
            df_multi = pd.read_excel(io.BytesIO(excel_bytes), header=[0, 1])
            df_flat, was_multi = column_mapper.flatten_multiheader(df_multi)
            if was_multi:
                df = df_flat
        except Exception:
            pass

    df.columns = [str(c).strip() for c in df.columns]

    # Rebuild year_groups with int keys
    year_groups = {}
    for yr_str, fields in year_groups_raw.items():
        year_groups[int(yr_str)] = fields
    mapping_result = {
        "mapping":    mapping,
        "year_groups": year_groups,
        "missing_required": [],
    }

    products, err = column_mapper.build_products_from_mapping(df, mapping_result)
    if err:
        return jsonify({"error": err}), 422

    adapter = da.InMemoryAdapter(products)
    session["adapter"]  = adapter
    session["meta"] = {
        "categories":  adapter.get_categories(),
        "attributes":  adapter.get_attributes(),
        "years":       adapter.get_years(),
        "suppliers":   adapter.get_suppliers(),
        "product_count": len({p["product_code"] for p in products}),
    }
    del session["_pending_excel_bytes"]

    return jsonify({
        "session_id":    session_id,
        "product_count": session["meta"]["product_count"],
        "categories":    session["meta"]["categories"],
        "attributes":    session["meta"]["attributes"],
        "years":         session["meta"]["years"],
    })


# ── Chat ──────────────────────────────────────────────────────────────────────

@app.route("/api/chat", methods=["POST"])
@auth.require_auth
def chat():
    body = request.get_json(silent=True) or {}
    session_id = str(body.get("session_id", ""))
    message    = str(body.get("message", "")).strip()
    history    = body.get("history", [])

    if not message:
        return jsonify({"error": "Messaggio vuoto."}), 400

    session = _sessions.get(session_id)
    if not session:
        return jsonify({"error": "Sessione non trovata. Ricarica il file Excel."}), 404
    if session.get("user_id") != g.user_id:
        return jsonify({"error": "Non autorizzato."}), 403

    adapter = session.get("adapter")
    if not adapter:
        return jsonify({"error": "Dati non ancora caricati. Completa il caricamento del file prima."}), 400

    meta = session["meta"]

    # Check if this is a clarification response (user picked an option)
    clarification_answer = body.get("clarification_answer")
    resolved_intent      = body.get("resolved_intent")

    if resolved_intent:
        # User confirmed a proposed intent — execute directly
        intent = resolved_intent
    elif clarification_answer and body.get("pending_intent"):
        intent = _apply_clarification(body["pending_intent"], clarification_answer, meta)
    else:
        intent = intent_engine.parse_intent(
            message=message,
            history=history,
            available_categories=meta["categories"],
            available_attributes=meta["attributes"],
            available_years=meta["years"],
            available_suppliers=meta.get("suppliers", []),
        )

    if intent.get("needs_clarification"):
        return jsonify({
            "type":           "clarification",
            "payload":        response_builder.build_clarification_response(intent, meta),
            "pending_intent": intent,
        })

    result  = query_engine.execute_query(intent, adapter)
    intent_type = intent.get("intent", "report")

    if intent_type == "report":
        payload = response_builder.build_report_response(result)
    elif intent_type == "planning":
        payload = response_builder.build_planning_response(result)
    else:
        payload = response_builder.build_both_response(result)

    return jsonify({"type": payload["type"], "payload": payload})


def _apply_clarification(pending_intent: dict, answer: str, meta: dict) -> dict:
    """Resolve a pending clarification with the user's answer."""
    ctype = pending_intent.get("clarification_type")

    if ctype == "multi_turn_confirm":
        if answer.lower().startswith("sì") or answer.lower().startswith("si") or answer == "Sì, procedi":
            proposed = pending_intent.get("proposed_filter", {})
            return {**pending_intent, **proposed, "needs_clarification": False}
        # User said no — return a new clarification asking them to rephrase
        return {
            **pending_intent,
            "needs_clarification": True,
            "clarification_type": None,
            "clarification_question": "Riscrivi la tua query con i filtri che vuoi applicare.",
        }

    if ctype == "category_ambiguous":
        candidates = pending_intent.get("candidate_categories", [])
        if answer == "Entrambe" or answer == "all":
            selected = candidates
        else:
            selected = [c for c in candidates if c in answer or answer in c]
            if not selected:
                selected = [answer] if answer in meta["categories"] else candidates
        intent = dict(pending_intent)
        intent["product_filter"]["categories"] = selected
        intent["needs_clarification"] = False
        return intent

    # For other types, just clear needs_clarification and let the query run
    return {**pending_intent, "needs_clarification": False}


# ── Export planning to Excel ──────────────────────────────────────────────────

@app.route("/api/export-planning", methods=["POST"])
@auth.require_auth
def export_planning():
    body  = request.get_json(silent=True) or {}
    cards = body.get("cards", [])
    if not cards:
        return jsonify({"error": "Nessun dato da esportare."}), 400

    status_labels = {
        "growing":      "In crescita",
        "stable":       "Stabile",
        "declining":    "In calo",
        "discontinued": "Non ordinare",
        "low_volume":   "Non ordinare",
    }
    rows = []
    for c in cards:
        rows.append({
            "Prodotto":             c.get("label", ""),
            "Stato":                status_labels.get(c.get("status", ""), c.get("status", "")),
            "Trend %":              c.get("trend_pct"),
            "Media storica (pz)":   c.get("avg_sales"),
            "Previsione (pz)":      c.get("forecast_demand"),
            "Stock attuale (pz)":   c.get("current_stock"),
            "Stock vendibile (pz)": c.get("effective_stock"),
            "Da ordinare (pz)":     c.get("recommended_purchase"),
            "Margine %":            c.get("gross_margin_pct"),
            "SL target %":          c.get("service_level_pct"),
            "Costo unitario (€)":   c.get("unit_cost"),
            "Prezzo vendita (€)":   c.get("unit_price"),
            "Margine atteso (€)":   c.get("margin_value"),
            "Metodologia":          c.get("methodology"),
            "Note AI":              c.get("ai_reasoning"),
        })

    df = pd.DataFrame(rows)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Pianificazione")
        ws = writer.sheets["Pianificazione"]
        # Basic column width
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
    buf.seek(0)

    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="pianificazione_acquisti.xlsx",
    )


# ── Health ────────────────────────────────────────────────────────────────────

@app.route("/api/health", methods=["GET"])
def health():
    return jsonify({"status": "ok"})


if __name__ == "__main__":
    debug = os.getenv("FLASK_DEBUG", "false").lower() == "true"
    port  = int(os.getenv("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=debug)
